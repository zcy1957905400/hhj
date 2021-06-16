# 接口自动化的执行步骤：
# 1、编写接口自动化的测试用例，读取excel里的测试数据。
# 2、写代码发送请求，得到接口的返回信息。
# 3、执行结果 vs 预期结果。
# 4、最终的实际结果写回到excel。
import json

import openpyxl
import requests
import jsonpath

# 读取测试用例数据
def read_data(filename,sheetname):
    workbook=openpyxl.load_workbook(filename) #加载工作簿
    sheet=workbook[sheetname] #获取sheet表单
    max_row=sheet.max_row #获取最大行数
    case_list=[] #定义一个空列表，存放测试用例
    for i in range(2,max_row+1): #+1取头不取尾
        dict_1=dict(case_id=sheet.cell(row=i,column=1).value, #用例编号
        url=sheet.cell(row=i, column=5).value,  #接口地址
        data=sheet.cell(row=i, column=6).value, #请求参数
        excepted=sheet.cell(row=i,column=7).value) #预期结果
        # print(dict_1)
        case_list.append(dict_1) #把字典数据一条一条追加到列表里存储
    return case_list

# 发送请求
def api_func(url,data):
    headers={"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
    res = requests.post(url=url, json=data, headers=headers)
    return res

#写入结果
def write_result(filename,sheetname,final_res,row,column):
    wb=openpyxl.load_workbook(filename) #加载工作薄
    sheet=wb[sheetname] #获取sheet表单
    sheet.cell(row=row,column=column).value=final_res #写入结果
    wb.save(filename) #保存测试用例

#执行自动化脚本
def execute_func(filename,sheetname):
    cases=read_data(filename,sheetname)  #调用读取函数
    for case in cases:  #从读取函数返回的数据进行取值
        case_id=int(case['case_id'])  #取出用例编号
        url=case.get('url') #取出接口地址
        data=case.get('data')  #取出请求参数，excel中读取的数据都是str类型
        data=eval(data)   #使用eval函数进行类型转换-->运行被字符串包裹的python表达式
        # print(data)
        excepted=case.get('excepted') #取出预期结果
        excepted=eval(excepted)
        excepted_msg=excepted.get('msg') #取出预期结果中的msg
        real_res=api_func(url,data).json() #调用发送请求的接口
        real_res_str=str(real_res)
        # print(real_res_str)
        real_res_msg=real_res.get('msg') #取出实际结果中的msg
        print('预期结果为：{}'.format(excepted_msg))
        print('实际结果为：{}'.format(real_res_msg))
        if excepted_msg==real_res_msg:
            print("这条用例通过")
            final_res='pass'
        else:
            print("这条用例不通过")
            final_res='fail'
        print('*'*30)
        write_result(filename, sheetname, real_res_str, case_id + 1, 8)
        write_result(filename, sheetname, final_res, case_id + 1, 9)

execute_func('test_case_api.xlsx', 'register')
execute_func('test_case_api.xlsx', 'login')

